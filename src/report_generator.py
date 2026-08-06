from pathlib import Path
from typing import Dict, Any, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, KeepTogether, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from src.config import REPORTS_DIR, CLEANED_DATA_DIR, CHARTS_DIR
from src.database import db_manager
from src.logger import logger

class ReportGenerator:
    """
    Report Exporter Engine.
    Generates cleaned CSVs, professionally formatted Excel reports (OpenPyXL),
    and PDF executive summary reports (ReportLab).
    """

    @staticmethod
    def export_cleaned_csv(df: pd.DataFrame, filename: str = "cleaned_sales_data.csv") -> Path:
        """Exports cleaned dataset as CSV."""
        output_path = CLEANED_DATA_DIR / filename
        df.to_csv(output_path, index=False)
        db_manager.log_report(filename, "CSV", str(output_path))
        logger.info(f"Cleaned CSV exported to: {output_path}")
        return output_path

    @staticmethod
    def generate_excel_report(df: pd.DataFrame, kpis: Dict[str, Any], monthly_df: pd.DataFrame, 
                              cat_df: pd.DataFrame, filename: str = "Sales_BI_Executive_Report.xlsx") -> Path:
        """
        Generates styled multi-tab Excel report using OpenPyXL.
        """
        output_path = REPORTS_DIR / filename
        wb = Workbook()

        # Sheet 1: Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.views.sheetView[0].showGridLines = True

        header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="1A237E")
        kpi_val_font = Font(name="Calibri", size=14, bold=True, color="00ACC1")
        kpi_lbl_font = Font(name="Calibri", size=10, bold=True, color="555555")

        thin_border = Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD")
        )

        ws1["A1"] = "BUSINESS INTELLIGENCE EXECUTIVE SALES SUMMARY"
        ws1["A1"].font = title_font
        ws1.merge_cells("A1:E1")

        # KPI Block Card Table
        kpi_cards = [
            ("Total Revenue", f"${kpis.get('total_revenue', 0):,.2f}"),
            ("Total Profit", f"${kpis.get('total_profit', 0):,.2f}"),
            ("Total Orders", f"{kpis.get('total_orders', 0):,}"),
            ("Avg Order Value", f"${kpis.get('avg_order_value', 0):,.2f}"),
            ("Profit Margin", f"{kpis.get('profit_margin', 0)}%")
        ]

        ws1.cell(row=3, column=1, value="Metric").font = header_font
        ws1.cell(row=3, column=1).fill = header_fill
        ws1.cell(row=3, column=2, value="Value").font = header_font
        ws1.cell(row=3, column=2).fill = header_fill

        for idx, (label, val) in enumerate(kpi_cards, start=4):
            c1 = ws1.cell(row=idx, column=1, value=label)
            c2 = ws1.cell(row=idx, column=2, value=val)
            c1.font = kpi_lbl_font
            c2.font = kpi_val_font
            c1.border = thin_border
            c2.border = thin_border

        # Category Table
        row_start = 11
        ws1.cell(row=row_start, column=1, value="Category Performance Breakdown").font = Font(size=13, bold=True, color="1A237E")
        
        cat_headers = ["Category", "Revenue ($)", "Profit ($)", "Profit Margin (%)"]
        for c_idx, h in enumerate(cat_headers, start=1):
            cell = ws1.cell(row=row_start+1, column=c_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill

        if not cat_df.empty:
            for r_idx, row in cat_df.iterrows():
                curr_row = row_start + 2 + r_idx
                ws1.cell(row=curr_row, column=1, value=row.get("category", "")).border = thin_border
                ws1.cell(row=curr_row, column=2, value=float(row.get("revenue", 0))).number_format = "$#,##0.00"
                ws1.cell(row=curr_row, column=3, value=float(row.get("profit", 0))).number_format = "$#,##0.00"
                ws1.cell(row=curr_row, column=4, value=float(row.get("profit_margin_%", 0))).number_format = "0.00'%' "
                for c in range(1, 5):
                    ws1.cell(row=curr_row, column=c).border = thin_border

        # Sheet 2: Raw Cleaned Dataset
        ws2 = wb.create_sheet(title="Cleaned Sales Data")
        ws2.views.sheetView[0].showGridLines = True

        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws2.cell(row=1, column=c_idx, value=col_name.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r_idx, row in df.iterrows():
            for c_idx, val in enumerate(row, start=1):
                if isinstance(val, pd.Timestamp):
                    cell = ws2.cell(row=r_idx+2, column=c_idx, value=val.strftime("%Y-%m-%d"))
                else:
                    cell = ws2.cell(row=r_idx+2, column=c_idx, value=val)
                cell.border = thin_border

        # Adjust Column Widths
        for sheet in [ws1, ws2]:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(output_path)
        db_manager.log_report(filename, "EXCEL", str(output_path))
        logger.info(f"Excel report generated at: {output_path}")
        return output_path

    @staticmethod
    def generate_pdf_report(df: pd.DataFrame, kpis: Dict[str, Any], insights: List[str], 
                            filename: str = "Sales_BI_Executive_Report.pdf") -> Path:
        """
        Generates Executive PDF Report using ReportLab.
        """
        output_path = REPORTS_DIR / filename
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=letter,
            rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
        )

        styles = getSampleStyleSheet()
        
        # Custom Styles
        title_style = ParagraphStyle(
            "DocTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#1A237E"),
            spaceAfter=4
        )
        
        subtitle_style = ParagraphStyle(
            "DocSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            textColor=colors.HexColor("#666666"),
            spaceAfter=15
        )

        section_heading = ParagraphStyle(
            "SectionHeading",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#1E88E5"),
            spaceBefore=12,
            spaceAfter=6
        )

        body_style = ParagraphStyle(
            "BodyTextCustom",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=13,
            textColor=colors.HexColor("#333333"),
            spaceAfter=6
        )

        elements = []

        # Title Banner
        elements.append(Paragraph("BUSINESS INTELLIGENCE SALES ANALYSIS REPORT", title_style))
        elements.append(Paragraph("Automated Executive Summary & Data Insights", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#1E88E5"), spaceAfter=12))

        # KPI Summary Table
        elements.append(Paragraph("Executive Key Performance Indicators", section_heading))
        kpi_data = [
            [
                Paragraph("<b>Total Revenue</b>", body_style),
                Paragraph(f"<font color='#1E88E5'><b>${kpis.get('total_revenue', 0):,.2f}</b></font>", body_style),
                Paragraph("<b>Total Profit</b>", body_style),
                Paragraph(f"<font color='#4CAF50'><b>${kpis.get('total_profit', 0):,.2f}</b></font>", body_style)
            ],
            [
                Paragraph("<b>Total Orders</b>", body_style),
                Paragraph(f"<b>{kpis.get('total_orders', 0):,}</b>", body_style),
                Paragraph("<b>Avg Order Value</b>", body_style),
                Paragraph(f"<b>${kpis.get('avg_order_value', 0):,.2f}</b>", body_style)
            ],
            [
                Paragraph("<b>Profit Margin</b>", body_style),
                Paragraph(f"<b>{kpis.get('profit_margin', 0)}%</b>", body_style),
                Paragraph("<b>Top Region</b>", body_style),
                Paragraph(f"<b>{kpis.get('top_region', 'N/A')}</b>", body_style)
            ]
        ]
        
        kpi_table = Table(kpi_data, colWidths=[120, 150, 120, 150])
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F6F9")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E0E0")),
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#1E88E5")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 10))

        # BI Insights Section
        elements.append(Paragraph("Automated Business Intelligence Insights", section_heading))
        for insight in insights:
            # Convert markdown bold to html bold for reportlab
            formatted_text = insight.replace("**", "<b>", 1).replace("**", "</b>", 1)
            elements.append(Paragraph(f"• {formatted_text}", body_style))

        elements.append(Spacer(1, 10))

        # Embedded Charts Section
        elements.append(Paragraph("Visual Sales Trends & Distribution", section_heading))
        
        revenue_chart = CHARTS_DIR / "revenue_trend.png"
        cat_chart = CHARTS_DIR / "category_sales.png"

        chart_elements = []
        if revenue_chart.exists():
            chart_elements.append(Image(str(revenue_chart), width=260, height=130))
        if cat_chart.exists():
            chart_elements.append(Image(str(cat_chart), width=260, height=130))

        if len(chart_elements) == 2:
            chart_table = Table([[chart_elements[0], chart_elements[1]]], colWidths=[270, 270])
            chart_table.setStyle(TableStyle([("ALIGN", (0,0), (-1,-1), "CENTER")]))
            elements.append(chart_table)
        elif len(chart_elements) == 1:
            elements.append(chart_elements[0])

        elements.append(Spacer(1, 10))

        # Top 5 Products Table
        prod_col = "product_name" if "product_name" in df.columns else df.columns[0]
        if prod_col in df.columns and not df.empty:
            elements.append(Paragraph("Top Performing Products", section_heading))
            top_df = df.groupby(prod_col).agg(
                rev=("sales", "sum"),
                prof=("profit", "sum")
            ).reset_index().sort_values(by="rev", ascending=False).head(5)

            table_data = [[
                Paragraph("<b>Product Name</b>", body_style),
                Paragraph("<b>Revenue ($)</b>", body_style),
                Paragraph("<b>Profit ($)</b>", body_style)
            ]]
            
            for _, r in top_df.iterrows():
                table_data.append([
                    Paragraph(str(r[prod_col]), body_style),
                    Paragraph(f"${r['rev']:,.2f}", body_style),
                    Paragraph(f"${r['prof']:,.2f}", body_style)
                ])

            prod_table = Table(table_data, colWidths=[300, 120, 120])
            prod_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E88E5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#1E88E5")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(prod_table)

        doc.build(elements)
        db_manager.log_report(filename, "PDF", str(output_path))
        logger.info(f"PDF report generated at: {output_path}")
        return output_path
